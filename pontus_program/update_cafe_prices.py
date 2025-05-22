import csv
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re

FILEPATH_OLD = "old.csv"
FILEPATH_UPD = "updates.pdf"
FILEPATH_NEW = "new.xlsx"

old = []
upd = {}
new = []

# Importing old price data
with open(FILEPATH_OLD, newline='', encoding='utf-8') as csvfile:
    reader = csv.DictReader(csvfile)
    for row in reader:
        if not row["Artikelnummer"]:
            continue
        
        entry_info = {
            "art_no" : int(row["Artikelnummer"]),
            "name" : row["Artikel"],
            "prev_price" : float(row["Inköpspris"].replace(',', '.')) if row["Inköpspris"] else 0.0,
            "new_price" : float(row["Inköpspris"].replace(',', '.')) if row["Inköpspris"] else 0.0,
            "updated" : False,
            "changed" : False,
            "comment" : None
        }
        
        old.append(entry_info)

# Importing new price data
line_pattern = re.compile(r"^(\d{4,6})\s+(.*?)\s+([A-ZÅÄÖ]{2,4})\s+(\d+,\d{2})$")

with pdfplumber.open(FILEPATH_UPD) as pdf:
    for page in pdf.pages:
        text = page.extract_text()

        for line in text.split("\n"):
            line = line.strip()
            match = line_pattern.match(line)
            if match:
                artnr = int(match.group(1))
                price = float(match.group(4).replace(",", "."))

                upd[artnr] = price
                
# Processing data
red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
orange = PatternFill(start_color="F1C232", end_color="F1C232", fill_type="solid")
green = PatternFill(start_color="9AD380", end_color="9AD380", fill_type="solid")

for art in old:
    no = art["art_no"]
    
    if not no in upd.keys():
        art["updated"] = False
        art["changed"] = False
        
        art["comment"] = "Article not found in list from Svensk Cater"
        
    else:
        art["updated"] = True
        art["new_price"] = upd[art["art_no"]]
        art["changed"] = art["new_price"] != art["prev_price"]
        
    
# Exporting data
wb = Workbook()
ws = wb.active
ws.title = "Updated prices"

ws.append(["Art. no", "Name", "Price", "Change", "Comment"])

for row, art in enumerate(old):
    artno = art["art_no"]
    name = art["name"]
    price = art["new_price"]
    
    change = None
    comment = None
    
    fac = 1.0
    
    if art["changed"] and art["prev_price"] != 0.0:
        fac = art["new_price"] / art["prev_price"]
        change = f"{(fac-1.)*100.0:.2f}%"
    
    if not art["updated"]:
        comment = art["comment"]
                        
    ws.append([artno, name, price, change, comment])
    
    if not art["updated"]:
            ws.cell(row+2, 1).fill = red
            ws.cell(row+2, 2).fill = red
            ws.cell(row+2, 4).fill = red
            ws.cell(row+2, 5).fill = red
    
    if fac > 1.0:
        ws.cell(row+2, 4).fill = orange
    elif fac < 1.0:
        ws.cell(row+2, 4).fill = green
    
wb.save(FILEPATH_NEW)